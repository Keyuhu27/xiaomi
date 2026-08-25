"""
把京东商智相关的 Excel 模板文件（如"数据源"sheet 里累积的历史数据）解析后
写入本地 DuckDB。

用法：
    python src/import_data.py                       # 处理 data/inbox/ 目录下所有文件
    python src/import_data.py path/to/file.xlsx      # 只处理指定文件

这个脚本不是按"文件"识别数据类型，而是打开工作簿后逐个 sheet 检查表头，
自动识别出 4 种已知的京东商智报表类型（sku_daily / funnel_daily /
keyword_brand_weekly / keyword_sku_rank_weekly，具体字段含义见 README 和
src/db.py 里的表结构注释），认不出的 sheet（汇总、透视表之类）会跳过。

日常使用流程：每天把京东商智下载的新数据粘贴进 Excel 模板对应的"数据源"
sheet 里（和你原来手工维护模板的习惯一样），然后把整份模板文件丢进
data/inbox/ 运行本脚本即可——重复导入同一天/同一系列/同一文件的数据会被
整体替换，不会产生重复行。
"""
import argparse
from datetime import timedelta
from pathlib import Path

import openpyxl
import pandas as pd

from db import get_connection
from parsing import parse_range_estimate, to_bool_cn, to_date, to_float, to_int

INBOX_DIR = Path(__file__).resolve().parent.parent / "data" / "inbox"
PROCESSED_DIR = INBOX_DIR / "processed"

# 每种报表用哪一列做"整体替换"的分区键：重新导入时，先删掉这一列取值相同
# 的旧数据，再整体插入新数据，这样同一份数据重复导入不会重复计数。
PARTITION_COL = {
    "funnel_daily": "series_code",       # 按产品系列整体替换
    "sku_daily": "sales_date",           # 按销售日期整体替换
    "keyword_brand_weekly": "source_file",      # 按来源文件整体替换
    "keyword_sku_rank_weekly": "source_file",
}

# 各指标的"成交子单量"等字段是整数、还是比率/金额这种浮点数
_FUNNEL_INT_METRICS = {
    "pv", "uv", "add_cart_customers", "follow_customers",
    "order_customers", "paying_customers", "sales_qty",
}
_FUNNEL_METRIC_FIELDS = {
    "浏览量(PV)": "pv",
    "访客数(UV)": "uv",
    "人均浏览量(人均PV)": "pv_per_uv",
    "平均停留时长(秒)": "avg_stay_seconds",
    "UV价值": "uv_value",
    "客单价": "aov",
    "加购客户数": "add_cart_customers",
    "关注客户数": "follow_customers",
    "加购转换率": "add_cart_rate",
    "下单客户数": "order_customers",
    "下单转换率": "order_rate",
    "下单金额": "order_amount",
    "成交客户数": "paying_customers",
    "成交转化率": "conversion_rate",
    "成交金额": "sales_amount",
    "成交子单量": "sales_qty",
}


def _raw(v):
    return None if v is None else str(v)


def detect_report_type(header: list) -> str | None:
    hs = {h for h in header if h is not None}
    if {"浏览量(PV)", "成交子单量", "一级渠道"} <= hs:
        return "funnel_daily"
    if {"SKU", "库存件数", "昨日出库商品件数", "一级类目"} <= hs:
        return "sku_daily"
    if {"关键词", "在线商品数", "搜索人数"} <= hs:
        return "keyword_brand_weekly"
    if {"SKUID", "商品信息", "排名"} <= hs:
        return "keyword_sku_rank_weekly"
    return None


def _funnel_metric_positions(header: list):
    """把每个指标列和紧跟其后的'同比'列配对，返回 [(db字段名, 指标列idx, 同比列idx或None), ...]。"""
    pairs = []
    for cn_name, field in _FUNNEL_METRIC_FIELDS.items():
        idx = header.index(cn_name)
        yoy_idx = idx + 1 if idx + 1 < len(header) and header[idx + 1] == "同比" else None
        pairs.append((field, idx, yoy_idx))
    return pairs


def load_funnel_daily(header, rows_iter, sheet_name, source_file):
    names = [
        "日期", "一级渠道", "二级渠道", "聚合来源", "核心节点", "周报周", "周会周",
        "是否近7天", "是否近14天", "预约期", "首销4小时", "首销28小时", "首销3天", "首销7天", "首销30天",
    ]
    try:
        idx = {name: header.index(name) for name in names}
        metric_positions = _funnel_metric_positions(header)
    except ValueError as e:
        print(f"  [跳过] sheet '{sheet_name}' 缺少必要列: {e}")
        return None

    series_code = sheet_name[len("数据源"):] if sheet_name.startswith("数据源") else sheet_name

    records = []
    for row in rows_iter:
        d = to_date(row[idx["日期"]])
        if d is None:
            continue
        rec = {
            "date": d,
            "series_code": series_code,
            "week_report": row[idx["周报周"]],
            "week_meeting": row[idx["周会周"]],
            "traffic_source": row[idx["聚合来源"]],
            "core_node": row[idx["核心节点"]],
            "is_last_7d": to_bool_cn(row[idx["是否近7天"]]),
            "is_last_14d": to_bool_cn(row[idx["是否近14天"]]),
            "is_reservation_period": to_bool_cn(row[idx["预约期"]]),
            "is_launch_4h": to_bool_cn(row[idx["首销4小时"]]),
            "is_launch_28h": to_bool_cn(row[idx["首销28小时"]]),
            "is_launch_3d": to_bool_cn(row[idx["首销3天"]]),
            "is_launch_7d": to_bool_cn(row[idx["首销7天"]]),
            "is_launch_30d": to_bool_cn(row[idx["首销30天"]]),
            "channel_l1": row[idx["一级渠道"]],
            "channel_l2": row[idx["二级渠道"]],
            "source_file": source_file,
        }
        for field, midx, yidx in metric_positions:
            raw = row[midx]
            rec[field] = to_int(raw) if field in _FUNNEL_INT_METRICS else to_float(raw)
            rec[f"{field}_yoy"] = to_float(row[yidx]) if yidx is not None else None
        records.append(rec)

    return pd.DataFrame.from_records(records) if records else None


def load_sku_daily(header, rows_iter, sheet_name, source_file):
    names = [
        "时间", "商品名称", "SKU", "品牌", "一级类目", "二级类目", "三级类目", "店铺名称",
        "RDC", "配送中心", "上下柜状态", "商品价格", "库存件数", "可用库存",
        "昨日出库商品件数", "近7日出库商品件数", "近14日出库商品件数", "近28日出库商品件数", "近30日出库商品件数",
    ]
    try:
        idx = {name: header.index(name) for name in names}
    except ValueError as e:
        print(f"  [跳过] sheet '{sheet_name}' 缺少必要列: {e}")
        return None

    records = []
    for row in rows_iter:
        snapshot_date = to_date(row[idx["时间"]])
        sku_id = row[idx["SKU"]]
        if snapshot_date is None or sku_id is None:
            continue
        records.append({
            "snapshot_date": snapshot_date,
            "sales_date": snapshot_date - timedelta(days=1),
            "sku_id": str(sku_id).strip(),
            "product_name": row[idx["商品名称"]],
            "brand": row[idx["品牌"]],
            "category_l1": row[idx["一级类目"]],
            "category_l2": row[idx["二级类目"]],
            "category_l3": row[idx["三级类目"]],
            "store_name": row[idx["店铺名称"]],
            "rdc": row[idx["RDC"]],
            "distribution_center": row[idx["配送中心"]],
            "shelf_status": row[idx["上下柜状态"]],
            "price": to_float(row[idx["商品价格"]]),
            "stock_qty": to_int(row[idx["库存件数"]]),
            "available_stock": to_int(row[idx["可用库存"]]),
            "sales_qty": to_int(row[idx["昨日出库商品件数"]]),
            "sales_qty_7d": to_int(row[idx["近7日出库商品件数"]]),
            "sales_qty_14d": to_int(row[idx["近14日出库商品件数"]]),
            "sales_qty_28d": to_int(row[idx["近28日出库商品件数"]]),
            "sales_qty_30d": to_int(row[idx["近30日出库商品件数"]]),
            "source_file": source_file,
        })

    return pd.DataFrame.from_records(records) if records else None


def load_keyword_brand_weekly(header, rows_iter, sheet_name, source_file):
    names = [
        "日期", "年", "月", "周", "SPU", "词条性质", "品牌", "子品牌", "是否近14天", "排名", "关键词",
        "搜索人数", "搜索次数", "点击人数", "点击次数", "点击率", "成交金额", "成交单量", "成交转化率", "在线商品数",
    ]
    try:
        idx = {name: header.index(name) for name in names}
    except ValueError as e:
        print(f"  [跳过] sheet '{sheet_name}' 缺少必要列: {e}")
        return None

    records = []
    for row in rows_iter:
        week_date = to_date(row[idx["日期"]])
        if week_date is None:
            continue
        records.append({
            "week_date": week_date,
            "year": row[idx["年"]],
            "month": row[idx["月"]],
            "week": row[idx["周"]],
            "spu": row[idx["SPU"]],
            "entry_type": row[idx["词条性质"]],
            "brand": row[idx["品牌"]],
            "sub_brand": row[idx["子品牌"]],
            "is_last_14d": to_bool_cn(row[idx["是否近14天"]]),
            "rank": to_int(row[idx["排名"]]),
            "keyword": row[idx["关键词"]],
            "search_users_raw": _raw(row[idx["搜索人数"]]), "search_users_est": parse_range_estimate(row[idx["搜索人数"]]),
            "search_count_raw": _raw(row[idx["搜索次数"]]), "search_count_est": parse_range_estimate(row[idx["搜索次数"]]),
            "click_users_raw": _raw(row[idx["点击人数"]]), "click_users_est": parse_range_estimate(row[idx["点击人数"]]),
            "click_count_raw": _raw(row[idx["点击次数"]]), "click_count_est": parse_range_estimate(row[idx["点击次数"]]),
            "click_rate_raw": _raw(row[idx["点击率"]]), "click_rate_est": parse_range_estimate(row[idx["点击率"]]),
            "sales_amount_raw": _raw(row[idx["成交金额"]]), "sales_amount_est": parse_range_estimate(row[idx["成交金额"]]),
            "sales_qty_raw": _raw(row[idx["成交单量"]]), "sales_qty_est": parse_range_estimate(row[idx["成交单量"]]),
            "conversion_rate_raw": _raw(row[idx["成交转化率"]]), "conversion_rate_est": parse_range_estimate(row[idx["成交转化率"]]),
            "online_sku_count": to_int(row[idx["在线商品数"]]),
            "source_file": source_file,
        })

    return pd.DataFrame.from_records(records) if records else None


def load_keyword_sku_rank_weekly(header, rows_iter, sheet_name, source_file):
    names = [
        "时间", "SPU", "是近14天", "品牌", "月", "周", "是否近14天", "排名",
        "SKUID", "商品信息", "品牌ID", "品牌名称", "点击人数", "点击次数", "成交单量", "成交金额",
    ]
    try:
        idx = {name: header.index(name) for name in names}
    except ValueError as e:
        print(f"  [跳过] sheet '{sheet_name}' 缺少必要列: {e}")
        return None

    records = []
    for row in rows_iter:
        week_date = to_date(row[idx["时间"]])
        if week_date is None:
            continue
        sku_id = row[idx["SKUID"]]
        brand_id = row[idx["品牌ID"]]
        records.append({
            "week_date": week_date,
            "spu": row[idx["SPU"]],
            "is_last_14d_1": to_bool_cn(row[idx["是近14天"]]),
            "brand": row[idx["品牌"]],
            "month": row[idx["月"]],
            "week": row[idx["周"]],
            "is_last_14d_2": to_bool_cn(row[idx["是否近14天"]]),
            "rank": to_int(row[idx["排名"]]),
            "sku_id": str(sku_id).strip() if sku_id is not None else None,
            "product_info": row[idx["商品信息"]],
            "brand_id": str(brand_id) if brand_id is not None else None,
            "brand_name": row[idx["品牌名称"]],
            "click_users_raw": _raw(row[idx["点击人数"]]), "click_users_est": parse_range_estimate(row[idx["点击人数"]]),
            "click_count_raw": _raw(row[idx["点击次数"]]), "click_count_est": parse_range_estimate(row[idx["点击次数"]]),
            "sales_qty_raw": _raw(row[idx["成交单量"]]), "sales_qty_est": parse_range_estimate(row[idx["成交单量"]]),
            "sales_amount_raw": _raw(row[idx["成交金额"]]), "sales_amount_est": parse_range_estimate(row[idx["成交金额"]]),
            "source_file": source_file,
        })

    return pd.DataFrame.from_records(records) if records else None


LOADERS = {
    "funnel_daily": load_funnel_daily,
    "sku_daily": load_sku_daily,
    "keyword_brand_weekly": load_keyword_brand_weekly,
    "keyword_sku_rank_weekly": load_keyword_sku_rank_weekly,
}


def write_table(con, table: str, df: pd.DataFrame, partition_col: str):
    for pval, group in df.groupby(partition_col):
        con.execute(f"DELETE FROM {table} WHERE {partition_col} = ?", [pval])
        con.register("tmp_df", group)
        cols = ", ".join(group.columns)
        con.execute(f"INSERT INTO {table} ({cols}) SELECT {cols} FROM tmp_df")
        con.unregister("tmp_df")


def import_workbook(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    con = get_connection()
    summary = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = list(next(rows_iter))
            except StopIteration:
                continue

            report_type = detect_report_type(header)
            if report_type is None:
                continue  # 汇总/透视表之类的辅助 sheet，跳过

            df = LOADERS[report_type](header, rows_iter, sheet_name, path.name)
            if df is None or df.empty:
                continue

            write_table(con, report_type, df, PARTITION_COL[report_type])
            summary.append((sheet_name, report_type, len(df)))
            print(f"  [完成] sheet '{sheet_name}' -> {report_type}: {len(df)} 行")
    finally:
        con.close()
        wb.close()

    if not summary:
        print(f"  [提示] {path.name} 里没有识别出任何已知的数据源 sheet（可能都是汇总/透视表，已跳过）")
    return summary


def _move_to_processed(path: Path):
    from datetime import datetime
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    target = PROCESSED_DIR / f"{datetime.now():%Y%m%d_%H%M%S}_{path.name}"
    path.rename(target)


def process_inbox():
    INBOX_DIR.mkdir(parents=True, exist_ok=True)
    files = [p for p in INBOX_DIR.iterdir() if p.is_file() and not p.name.startswith(".")]
    if not files:
        print(f"[提示] {INBOX_DIR} 里没有待导入的文件")
        return

    for path in files:
        print(f"[处理] {path.name}")
        try:
            import_workbook(path)
            _move_to_processed(path)
        except Exception as e:
            print(f"[失败] {path.name}: {e}")


def main():
    parser = argparse.ArgumentParser(description="导入京东商智相关 Excel 模板数据")
    parser.add_argument("file", nargs="?", help="指定单个文件；不传则处理 data/inbox/ 下所有文件")
    args = parser.parse_args()

    if args.file:
        print(f"[处理] {args.file}")
        import_workbook(Path(args.file))
    else:
        process_inbox()


if __name__ == "__main__":
    main()
